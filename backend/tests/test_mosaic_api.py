"""
Mosaic Financial Modeling API Tests
Tests all key API endpoints for the stock analysis app
"""
import pytest
import requests
import os
import time

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', 'https://stock-analysis-ai-1.preview.emergentagent.com').rstrip('/')


class TestHealthAndBasics:
    """Test basic health endpoints"""
    
    def test_api_root(self):
        """Test API root endpoint"""
        response = requests.get(f"{BASE_URL}/api/")
        assert response.status_code == 200
        data = response.json()
        assert "message" in data or "status" in data
        print(f"API root response: {data}")
    
    def test_jobs_list(self):
        """Test listing jobs endpoint"""
        response = requests.get(f"{BASE_URL}/api/generate/jobs")
        assert response.status_code == 200
        data = response.json()
        assert isinstance(data, list)
        print(f"Found {len(data)} jobs")
        
        # Verify job structure if there are jobs
        if len(data) > 0:
            job = data[0]
            assert "id" in job
            assert "ticker" in job
            assert "status" in job


class TestTickerValidation:
    """Test ticker validation endpoint"""
    
    def test_validate_valid_ticker(self):
        """Test validation of a valid ticker"""
        response = requests.get(f"{BASE_URL}/api/generate/validate-ticker/HDFCBANK")
        assert response.status_code == 200
        data = response.json()
        assert data["ticker"] == "HDFCBANK"
        assert data["valid"] is True
        print(f"Ticker validation: {data}")
    
    def test_validate_another_ticker(self):
        """Test validation of RELIANCE ticker"""
        response = requests.get(f"{BASE_URL}/api/generate/validate-ticker/RELIANCE")
        assert response.status_code == 200
        data = response.json()
        assert data["ticker"] == "RELIANCE"
        assert "valid" in data
        print(f"RELIANCE validation: {data}")
    
    def test_validate_invalid_ticker(self):
        """Test validation of an invalid ticker"""
        response = requests.get(f"{BASE_URL}/api/generate/validate-ticker/INVALIDXYZ123")
        assert response.status_code == 200
        data = response.json()
        assert "valid" in data
        print(f"Invalid ticker validation: {data}")


class TestJobCreation:
    """Test job creation and progress endpoints"""
    
    def test_create_job_returns_job(self):
        """Test creating a new job returns valid job object"""
        response = requests.post(
            f"{BASE_URL}/api/generate/",
            json={"ticker": "ICICIBANK"},
            headers={"Content-Type": "application/json"}
        )
        assert response.status_code == 200
        data = response.json()
        
        assert "id" in data
        assert data["ticker"] == "ICICIBANK"
        assert "status" in data
        print(f"Job created/found: {data['id']}, status: {data['status']}")
        
        return data["id"]


class TestJobProgress:
    """Test job progress endpoints"""
    
    def test_get_completed_job_progress(self):
        """Test getting progress of a known completed job"""
        # First get a completed job from the list
        response = requests.get(f"{BASE_URL}/api/generate/jobs")
        assert response.status_code == 200
        jobs = response.json()
        
        completed_jobs = [j for j in jobs if j.get("status") == "completed"]
        if not completed_jobs:
            pytest.skip("No completed jobs to test")
        
        job_id = completed_jobs[0]["id"]
        
        progress_response = requests.get(f"{BASE_URL}/api/generate/progress/{job_id}")
        assert progress_response.status_code == 200
        data = progress_response.json()
        
        assert data["status"] == "completed"
        assert "ticker" in data
        print(f"Completed job progress: {data['ticker']}, step: {data.get('current_step')}")


class TestJobResults:
    """Test job result endpoints"""
    
    def test_get_completed_job_result(self):
        """Test getting result of a completed job"""
        # First get a completed job
        response = requests.get(f"{BASE_URL}/api/generate/jobs")
        jobs = response.json()
        
        completed_jobs = [j for j in jobs if j.get("status") == "completed"]
        if not completed_jobs:
            pytest.skip("No completed jobs to test")
        
        job_id = completed_jobs[0]["id"]
        
        result_response = requests.get(f"{BASE_URL}/api/generate/result/{job_id}")
        assert result_response.status_code == 200
        data = result_response.json()
        
        assert "ticker" in data
        assert "result" in data
        
        # Verify result structure
        result = data["result"]
        assert "valuation" in result or "thesis" in result
        print(f"Result for {data['ticker']}: has valuation={bool(result.get('valuation'))}, thesis={bool(result.get('thesis'))}")
    
    def test_result_has_thesis_content(self):
        """Test that completed job result has actual thesis content"""
        response = requests.get(f"{BASE_URL}/api/generate/jobs")
        jobs = response.json()
        
        completed_jobs = [j for j in jobs if j.get("status") == "completed"]
        if not completed_jobs:
            pytest.skip("No completed jobs to test")
        
        job_id = completed_jobs[0]["id"]
        
        result_response = requests.get(f"{BASE_URL}/api/generate/result/{job_id}")
        assert result_response.status_code == 200
        data = result_response.json()
        
        result = data["result"]
        thesis = result.get("thesis", {})
        
        # Thesis should have actual content, not placeholder
        assert thesis.get("summary") is not None or thesis.get("bull_case") is not None, "Thesis should have content"
        
        # Check for bull_case and bear_case
        if thesis.get("bull_case"):
            assert len(thesis["bull_case"]) > 0, "Bull case should have items"
            print(f"Bull case has {len(thesis['bull_case'])} points")
        
        if thesis.get("bear_case"):
            assert len(thesis["bear_case"]) > 0, "Bear case should have items"
            print(f"Bear case has {len(thesis['bear_case'])} points")


class TestExcelDownload:
    """Test Excel download functionality"""
    
    def test_download_excel_returns_file(self):
        """Test downloading Excel file from a completed job"""
        # Get a completed job
        response = requests.get(f"{BASE_URL}/api/generate/jobs")
        jobs = response.json()
        
        completed_jobs = [j for j in jobs if j.get("status") == "completed" and j.get("excel_path")]
        if not completed_jobs:
            pytest.skip("No completed jobs with Excel to test")
        
        job_id = completed_jobs[0]["id"]
        
        download_response = requests.get(f"{BASE_URL}/api/generate/download/{job_id}")
        assert download_response.status_code == 200
        
        # Verify it's an Excel file
        content_type = download_response.headers.get("content-type", "")
        assert "spreadsheet" in content_type or len(download_response.content) > 0
        
        # Verify content is not empty
        assert len(download_response.content) > 1000, "Excel file should not be empty"
        print(f"Downloaded Excel file: {len(download_response.content)} bytes")
    
    def test_excel_is_valid_xlsx(self):
        """Test that downloaded Excel is a valid XLSX file"""
        import tempfile
        
        response = requests.get(f"{BASE_URL}/api/generate/jobs")
        jobs = response.json()
        
        completed_jobs = [j for j in jobs if j.get("status") == "completed" and j.get("excel_path")]
        if not completed_jobs:
            pytest.skip("No completed jobs with Excel to test")
        
        job_id = completed_jobs[0]["id"]
        
        download_response = requests.get(f"{BASE_URL}/api/generate/download/{job_id}")
        assert download_response.status_code == 200
        
        # Save to temp file and verify
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(download_response.content)
            temp_path = f.name
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(temp_path)
            sheets = wb.sheetnames
            
            assert len(sheets) >= 5, f"Excel should have at least 5 sheets, found {len(sheets)}"
            print(f"Excel sheets: {sheets}")
            
            # Check for expected sheets
            expected_sheets = ["Cover", "P&L", "Balance Sheet"]
            for sheet in expected_sheets:
                assert sheet in sheets, f"Missing expected sheet: {sheet}"
        finally:
            import os
            os.unlink(temp_path)


class TestCacheEndpoints:
    """Test cache-related endpoints"""
    
    def test_cache_status(self):
        """Test getting cache status for a ticker"""
        response = requests.get(f"{BASE_URL}/api/generate/cache-status/HDFCBANK")
        assert response.status_code == 200
        data = response.json()
        assert "has_cache" in data or "cached_steps" in data
        print(f"Cache status for HDFCBANK: {data}")
    
    def test_cached_tickers_list(self):
        """Test listing cached tickers"""
        response = requests.get(f"{BASE_URL}/api/generate/cached-tickers")
        assert response.status_code == 200
        data = response.json()
        assert "tickers" in data
        print(f"Cached tickers: {data['tickers'][:5] if data['tickers'] else 'none'}")


class TestErrorHandling:
    """Test error handling"""
    
    def test_job_not_found(self):
        """Test getting progress for non-existent job"""
        response = requests.get(f"{BASE_URL}/api/generate/progress/non-existent-id-12345")
        assert response.status_code == 404
    
    def test_result_not_found(self):
        """Test getting result for non-existent job"""
        response = requests.get(f"{BASE_URL}/api/generate/result/non-existent-id-12345")
        assert response.status_code == 404


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
