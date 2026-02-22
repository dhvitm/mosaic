from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
import logging
from typing import Dict, Any, List, Tuple
from datetime import datetime

logger = logging.getLogger(__name__)

# Define styles
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(size=16, bold=True, color="1F4E79")
SECTION_FONT = Font(size=12, bold=True, color="2F5496")
NUMBER_FONT = Font(name="Consolas", size=10)
FORECAST_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
ASSUMPTION_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
ROE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

class ExcelGenerator:
    def __init__(self):
        self.wb = None
        self.forecast_years = ['FY26E', 'FY27E', 'FY28E', 'FY29E', 'FY30E']
        # Store row references for formula linking
        self.pnl_rows = {}
        self.bs_rows = {}
        self.assumption_refs = {}
        
    def generate_model(self, job_id: str, data: Dict[str, Any]) -> str:
        """Generate complete Excel financial model with formulas"""
        try:
            self.wb = Workbook()
            
            if 'Sheet' in self.wb.sheetnames:
                self.wb.remove(self.wb['Sheet'])
            
            ticker = data.get('company_metadata', {}).get('ticker', 'UNKNOWN')
            company_name = data.get('company_metadata', {}).get('full_name', 'Unknown Company')
            
            # Create sheets in order - Assumptions first for formula references
            self._create_cover_sheet(company_name, ticker, data)
            self._create_assumptions_sheet(data)  # Must be first for formula refs
            self._create_pnl_with_formulas(data)
            self._create_balance_sheet_with_formulas(data)
            self._create_roe_tree_sheet(data)  # NEW: ROE decomposition
            self._create_quarterly_sheet(data)
            self._create_ratios_sheet(data)
            self._create_valuation_sheet(data)
            self._create_peer_comparison_sheet(data)
            self._create_thesis_sheet(data)
            
            # Save file
            import os
            os.makedirs("/app/generated_models", exist_ok=True)
            filename = f"{ticker}_model_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            filepath = f"/app/generated_models/{filename}"
            self.wb.save(filepath)
            
            logger.info(f"Excel model generated: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error generating Excel model: {str(e)}")
            raise
    
    def _apply_header_style(self, ws, row: int, start_col: int, end_col: int):
        """Apply header style to a row"""
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = THIN_BORDER
    
    def _get_historical_years(self, data: Dict) -> List[str]:
        """Extract historical years from scraped data"""
        historical = data.get('historical_financials', {})
        annual_pnl = historical.get('annual_pnl', {})
        if annual_pnl:
            years = sorted([y for y in annual_pnl.keys() if y != 'TTM'], key=lambda x: x)
            return years[-5:]  # Last 5 years
        return ['Mar 2021', 'Mar 2022', 'Mar 2023', 'Mar 2024', 'Mar 2025']
    
    def _create_cover_sheet(self, company_name: str, ticker: str, data: Dict[str, Any]):
        """Create cover sheet with summary"""
        ws = self.wb.create_sheet("Cover", 0)
        
        metadata = data.get('company_metadata', {})
        valuation = data.get('valuation', {})
        thesis = data.get('thesis', {})
        
        ws['A1'] = "FINANCIAL MODEL"
        ws['A1'].font = Font(size=24, bold=True, color="1F4E79")
        
        ws['A2'] = company_name
        ws['A2'].font = Font(size=18, bold=True)
        
        ws['A3'] = f"Ticker: {ticker}"
        ws['A3'].font = Font(size=12, color="666666")
        
        ws['A5'] = f"Generated: {datetime.now().strftime('%d %B %Y %H:%M')}"
        ws['A5'].font = Font(size=10, italic=True, color="999999")
        
        # Key Data
        ws['A7'] = "KEY DATA"
        ws['A7'].font = SECTION_FONT
        
        key_data = [
            ("Sector", metadata.get('sector', 'N/A')),
            ("Market Cap", f"Rs.{metadata.get('market_cap', 0):,.0f} Cr"),
            ("Current Price", f"Rs.{metadata.get('current_price', 0):,.2f}"),
        ]
        
        for i, (label, value) in enumerate(key_data, start=8):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'B{i}'].font = Font(bold=True)
        
        # Recommendation
        ws['A13'] = "RECOMMENDATION"
        ws['A13'].font = SECTION_FONT
        
        rec = valuation.get('recommendation', 'HOLD')
        rec_color = "00AA00" if rec == "BUY" else "CC0000" if rec == "SELL" else "FF9900"
        
        ws['A14'] = "Rating"
        ws['B14'] = rec
        ws['B14'].font = Font(size=14, bold=True, color=rec_color)
        
        ws['A15'] = "Target Price"
        ws['B15'] = f"Rs.{valuation.get('target_price', 0):,.0f}"
        ws['B15'].font = Font(size=14, bold=True)
        
        ws['A16'] = "Upside"
        ws['B16'] = f"{valuation.get('upside_percent', 0):.1f}%"
        
        # Thesis Summary
        ws['A19'] = "INVESTMENT THESIS"
        ws['A19'].font = SECTION_FONT
        ws['A20'] = thesis.get('summary', f"{rec} recommendation")
        ws['A20'].alignment = Alignment(wrap_text=True)
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def _create_assumptions_sheet(self, data: Dict[str, Any]):
        """Create Assumptions sheet - this is the driver sheet for forecasts"""
        ws = self.wb.create_sheet("Assumptions")
        
        assumptions_data = data.get('assumptions', {})
        assumptions = assumptions_data.get('assumptions', {})
        
        ws['A1'] = "FORECAST ASSUMPTIONS"
        ws['A1'].font = TITLE_FONT
        
        ws['A2'] = "Change yellow cells to update projections. All linked sheets will recalculate."
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        
        # Headers
        ws['A4'] = "Assumption"
        for i, year in enumerate(self.forecast_years, start=2):
            ws.cell(row=4, column=i).value = year
        self._apply_header_style(ws, 4, 1, len(self.forecast_years) + 1)
        
        # Define assumptions to show - store cell references for formulas
        assumption_rows = [
            ("revenue_growth_rate", "Revenue Growth Rate (%)", 100),
            ("loan_growth_rate", "Loan/Advances Growth (%)", 100),
            ("nim", "Net Interest Margin (%)", 100),
            ("ebitda_margin", "EBITDA Margin (%)", 100),
            ("operating_expense_growth", "Operating Expense Growth (%)", 100),
            ("depreciation_rate", "Depreciation Rate (%)", 100),
            ("interest_expense_rate", "Interest Expense Rate (%)", 100),
            ("tax_rate", "Effective Tax Rate (%)", 100),
            ("dividend_payout", "Dividend Payout (%)", 100),
            ("roa", "Return on Assets (%)", 100),
            ("roe", "Return on Equity (%)", 100),
            ("casa_ratio", "CASA Ratio (%)", 100),
            ("credit_cost", "Credit Cost (%)", 100),
            ("cost_to_income", "Cost-to-Income (%)", 100),
        ]
        
        row = 5
        self.assumption_refs = {}  # Reset for this generation
        
        for key, label, multiplier in assumption_rows:
            if key in assumptions:
                ws.cell(row=row, column=1).value = label
                ws.cell(row=row, column=1).border = THIN_BORDER
                
                values = assumptions[key]
                if isinstance(values, list):
                    for i, val in enumerate(values[:5], start=2):
                        cell = ws.cell(row=row, column=i)
                        cell.value = round(val * multiplier, 2) if val else 0
                        cell.number_format = '0.00'
                        cell.font = NUMBER_FONT
                        cell.border = THIN_BORDER
                        cell.fill = ASSUMPTION_FILL  # Yellow for editable
                        cell.alignment = Alignment(horizontal='right')
                        
                        # Store cell reference for formulas
                        year_idx = i - 2  # 0-indexed
                        if year_idx < len(self.forecast_years):
                            if key not in self.assumption_refs:
                                self.assumption_refs[key] = {}
                            self.assumption_refs[key][self.forecast_years[year_idx]] = f"Assumptions!${get_column_letter(i)}${row}"
                row += 1
        
        # Add note about defaults if assumptions are sparse
        row += 2
        ws.cell(row=row, column=1).value = "RATIONALE"
        ws.cell(row=row, column=1).font = SECTION_FONT
        ws.cell(row=row+1, column=1).value = assumptions_data.get('rationale', 'Based on historical trends and sector outlook')
        ws.cell(row=row+1, column=1).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=6)
        
        ws.column_dimensions['A'].width = 30
        for i in range(2, 8):
            ws.column_dimensions[get_column_letter(i)].width = 12
    
    def _create_pnl_with_formulas(self, data: Dict[str, Any]):
        """Create P&L with mechanical linking formulas and projections - using actual Screener keys"""
        ws = self.wb.create_sheet("P&L")
        
        ws['A1'] = "PROFIT & LOSS STATEMENT"
        ws['A1'].font = TITLE_FONT
        
        ws['A2'] = "(Rs. Crores) | Green cells = Forecast with formulas"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        
        # Get historical data
        historical = data.get('historical_financials', {})
        annual_pnl = historical.get('annual_pnl', {})
        
        hist_years = self._get_historical_years(data)
        all_years = hist_years + self.forecast_years
        
        # Headers
        ws['A4'] = "Line Item"
        for i, year in enumerate(all_years, start=2):
            cell = ws.cell(row=4, column=i)
            cell.value = year
            if year in self.forecast_years:
                cell.fill = FORECAST_FILL
        self._apply_header_style(ws, 4, 1, len(all_years) + 1)
        
        # P&L structure matching ACTUAL Screener.in keys for banks
        # Banks have simpler P&L: Revenue +, Interest, Expenses +, Other Income +, Depreciation, Profit before tax, Tax %, Net Profit +
        pnl_items = [
            ('Revenue / Interest Earned', ['Revenue +', 'Interest Earned', 'Revenue', 'Sales +']),
            ('Interest Income', ['Interest', 'Interest Income']),
            ('Other Income', ['Other Income +', 'Other Income']),
            ('Total Income', None),  # Calculated = sum of above
            ('', None),  # Spacer
            ('Operating Expenses', ['Expenses +', 'Operating Expenses', 'Expenses']),
            ('Depreciation', ['Depreciation', 'Depreciation and Amortisation']),
            ('Total Expenses', None),  # Calculated
            ('', None),  # Spacer
            ('Profit Before Tax', ['Profit before tax', 'PBT']),
            ('Tax', ['Tax %', 'Tax']),
            ('Net Profit', ['Net Profit +', 'Net Profit', 'PAT']),
            ('EPS (Rs.)', ['EPS in Rs', 'EPS']),
            ('Dividend Payout (%)', ['Dividend Payout %']),
        ]
        
        row = 5
        forecast_col_start = len(hist_years) + 2
        self.pnl_rows = {}
        
        for display_name, screener_keys in pnl_items:
            if not display_name:  # Spacer row
                row += 1
                continue
            
            ws.cell(row=row, column=1).value = display_name
            ws.cell(row=row, column=1).border = THIN_BORDER
            
            # Bold key items
            if 'Total' in display_name or 'Net Profit' in display_name or 'Profit Before' in display_name:
                ws.cell(row=row, column=1).font = Font(bold=True)
            
            # Store row reference for formulas
            key = display_name.lower().replace(' ', '_').replace('/', '_')
            self.pnl_rows[key] = row
            
            # Fill historical data
            if screener_keys and annual_pnl:
                for col_idx, year in enumerate(hist_years, start=2):
                    year_data = annual_pnl.get(year, {})
                    
                    # Try each possible key
                    value = None
                    for key in screener_keys:
                        if key in year_data:
                            value = year_data[key]
                            break
                    
                    cell = ws.cell(row=row, column=col_idx)
                    if value is not None:
                        cell.value = value
                        cell.number_format = '#,##0.00'
                    cell.border = THIN_BORDER
                    cell.font = NUMBER_FONT
                    cell.alignment = Alignment(horizontal='right')
            
            # Fill calculated fields for historical years
            elif display_name == 'Total Income' and annual_pnl:
                rev_row = self.pnl_rows.get('revenue___interest_earned', row-3)
                int_row = self.pnl_rows.get('interest_income', row-2)
                oth_row = self.pnl_rows.get('other_income', row-1)
                
                for col_idx, year in enumerate(hist_years, start=2):
                    col = get_column_letter(col_idx)
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = f"={col}{rev_row}+{col}{int_row}+{col}{oth_row}"
                    cell.number_format = '#,##0.00'
                    cell.border = THIN_BORDER
                    cell.font = Font(bold=True)
            
            elif display_name == 'Total Expenses' and annual_pnl:
                opex_row = self.pnl_rows.get('operating_expenses', row-2)
                dep_row = self.pnl_rows.get('depreciation', row-1)
                
                for col_idx, year in enumerate(hist_years, start=2):
                    col = get_column_letter(col_idx)
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = f"={col}{opex_row}+{col}{dep_row}"
                    cell.number_format = '#,##0.00'
                    cell.border = THIN_BORDER
                    cell.font = Font(bold=True)
            
            # Add formulas for forecast columns
            for fc_idx, fc_year in enumerate(self.forecast_years):
                col_idx = forecast_col_start + fc_idx
                cell = ws.cell(row=row, column=col_idx)
                prev_col = get_column_letter(col_idx - 1)
                curr_col = get_column_letter(col_idx)
                
                if display_name == 'Revenue / Interest Earned':
                    growth_ref = self.assumption_refs.get('revenue_growth_rate', {}).get(fc_year)
                    if growth_ref:
                        cell.value = f"={prev_col}{row}*(1+{growth_ref}/100)"
                    else:
                        cell.value = f"={prev_col}{row}*1.12"
                        
                elif display_name == 'Interest Income':
                    cell.value = f"={prev_col}{row}*1.10"
                    
                elif display_name == 'Other Income':
                    cell.value = f"={prev_col}{row}*1.08"
                    
                elif display_name == 'Total Income':
                    rev_row = self.pnl_rows.get('revenue___interest_earned')
                    int_row = self.pnl_rows.get('interest_income')
                    oth_row = self.pnl_rows.get('other_income')
                    cell.value = f"={curr_col}{rev_row}+{curr_col}{int_row}+{curr_col}{oth_row}"
                    cell.font = Font(bold=True)
                    
                elif display_name == 'Operating Expenses':
                    cell.value = f"={prev_col}{row}*1.10"
                    
                elif display_name == 'Depreciation':
                    cell.value = f"={prev_col}{row}*1.05"
                    
                elif display_name == 'Total Expenses':
                    opex_row = self.pnl_rows.get('operating_expenses')
                    dep_row = self.pnl_rows.get('depreciation')
                    cell.value = f"={curr_col}{opex_row}+{curr_col}{dep_row}"
                    cell.font = Font(bold=True)
                    
                elif display_name == 'Profit Before Tax':
                    inc_row = self.pnl_rows.get('total_income')
                    exp_row = self.pnl_rows.get('total_expenses')
                    cell.value = f"={curr_col}{inc_row}-{curr_col}{exp_row}"
                    cell.font = Font(bold=True)
                    
                elif display_name == 'Tax':
                    pbt_row = self.pnl_rows.get('profit_before_tax')
                    tax_ref = self.assumption_refs.get('tax_rate', {}).get(fc_year)
                    if tax_ref:
                        cell.value = f"={curr_col}{pbt_row}*{tax_ref}/100"
                    else:
                        cell.value = f"={curr_col}{pbt_row}*0.25"
                        
                elif display_name == 'Net Profit':
                    pbt_row = self.pnl_rows.get('profit_before_tax')
                    tax_row = self.pnl_rows.get('tax')
                    cell.value = f"={curr_col}{pbt_row}-{curr_col}{tax_row}"
                    cell.font = Font(bold=True)
                    
                elif display_name == 'EPS (Rs.)':
                    pat_row = self.pnl_rows.get('net_profit')
                    # EPS growth with PAT
                    cell.value = f"={prev_col}{row}*({curr_col}{pat_row}/{prev_col}{pat_row})"
                    
                elif display_name == 'Dividend Payout (%)':
                    cell.value = f"={prev_col}{row}"  # Keep constant
                
                cell.number_format = '#,##0.00'
                cell.border = THIN_BORDER
                cell.fill = FORECAST_FILL
                cell.alignment = Alignment(horizontal='right')
            
            row += 1
        
        # Add growth rate row
        row += 1
        ws.cell(row=row, column=1).value = "Net Profit Growth (%)"
        ws.cell(row=row, column=1).font = Font(bold=True, italic=True)
        pat_row = self.pnl_rows.get('net_profit')
        if pat_row:
            for col_idx in range(3, len(all_years) + 2):
                cell = ws.cell(row=row, column=col_idx)
                prev_col = get_column_letter(col_idx - 1)
                curr_col = get_column_letter(col_idx)
                cell.value = f"=IF({prev_col}{pat_row}=0,0,({curr_col}{pat_row}/{prev_col}{pat_row}-1)*100)"
                cell.number_format = '0.0'
                cell.border = THIN_BORDER
                if col_idx >= forecast_col_start:
                    cell.fill = FORECAST_FILL
        
        ws.column_dimensions['A'].width = 28
        for i in range(2, len(all_years) + 2):
            ws.column_dimensions[get_column_letter(i)].width = 14
    
    def _create_balance_sheet_with_formulas(self, data: Dict[str, Any]):
        """Create Balance Sheet with forecast projections - using actual Screener keys"""
        ws = self.wb.create_sheet("Balance Sheet")
        
        ws['A1'] = "BALANCE SHEET"
        ws['A1'].font = TITLE_FONT
        
        ws['A2'] = "(Rs. Crores) | Green cells = Forecast"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        
        historical = data.get('historical_financials', {})
        annual_bs = historical.get('annual_bs', {})
        
        hist_years = self._get_historical_years(data)
        all_years = hist_years + self.forecast_years
        
        # Headers
        ws['A4'] = "Line Item"
        for i, year in enumerate(all_years, start=2):
            cell = ws.cell(row=4, column=i)
            cell.value = year
            if year in self.forecast_years:
                cell.fill = FORECAST_FILL
        self._apply_header_style(ws, 4, 1, len(all_years) + 1)
        
        # Balance Sheet items matching ACTUAL Screener.in keys for banks
        # Screener keys: Equity Capital, Reserves, Deposits, Borrowing, Other Liabilities +, Total Liabilities,
        #                Fixed Assets +, CWIP, Investments, Other Assets +, Total Assets
        bs_items = [
            ('LIABILITIES', None),
            ('Equity Capital', ['Equity Capital', 'Share Capital']),
            ('Reserves', ['Reserves', 'Reserves and Surplus']),
            ('Total Equity', None),  # Calculated
            ('', None),  # Spacer
            ('Deposits', ['Deposits']),
            ('Borrowings', ['Borrowing', 'Borrowings']),
            ('Other Liabilities', ['Other Liabilities +', 'Other Liabilities']),
            ('Total Liabilities', ['Total Liabilities']),
            ('', None),  # Spacer
            ('ASSETS', None),
            ('Fixed Assets', ['Fixed Assets +', 'Fixed Assets']),
            ('CWIP', ['CWIP']),
            ('Investments', ['Investments']),
            ('Advances', ['Advances +', 'Loans and Advances']),
            ('Other Assets', ['Other Assets +', 'Other Assets']),
            ('Total Assets', ['Total Assets']),
        ]
        
        row = 5
        self.bs_rows = {}
        forecast_col_start = len(hist_years) + 2
        
        for display_name, screener_keys in bs_items:
            if not display_name:  # Spacer row
                row += 1
                continue
            
            ws.cell(row=row, column=1).value = display_name
            ws.cell(row=row, column=1).border = THIN_BORDER
            
            # Bold section headers and totals
            if display_name in ['LIABILITIES', 'ASSETS'] or 'Total' in display_name:
                ws.cell(row=row, column=1).font = Font(bold=True)
            
            # Store row reference
            key = display_name.lower().replace(' ', '_')
            self.bs_rows[key] = row
            
            # Fill historical data
            if screener_keys and annual_bs:
                for col_idx, year in enumerate(hist_years, start=2):
                    year_data = annual_bs.get(year, {})
                    
                    value = None
                    for key in screener_keys:
                        if key in year_data:
                            value = year_data[key]
                            break
                    
                    cell = ws.cell(row=row, column=col_idx)
                    if value is not None:
                        cell.value = value
                        cell.number_format = '#,##0.00'
                    cell.border = THIN_BORDER
                    cell.font = NUMBER_FONT
                    cell.alignment = Alignment(horizontal='right')
            
            # Fill calculated fields for historical years
            elif display_name == 'Total Equity' and annual_bs:
                eq_row = self.bs_rows.get('equity_capital')
                res_row = self.bs_rows.get('reserves')
                
                for col_idx, year in enumerate(hist_years, start=2):
                    col = get_column_letter(col_idx)
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = f"={col}{eq_row}+{col}{res_row}"
                    cell.number_format = '#,##0.00'
                    cell.border = THIN_BORDER
                    cell.font = Font(bold=True)
            
            # Add formulas for forecast columns
            for fc_idx, fc_year in enumerate(self.forecast_years):
                col_idx = forecast_col_start + fc_idx
                cell = ws.cell(row=row, column=col_idx)
                prev_col = get_column_letter(col_idx - 1)
                curr_col = get_column_letter(col_idx)
                
                if display_name == 'Equity Capital':
                    cell.value = f"={prev_col}{row}"  # Stays constant
                    
                elif display_name == 'Reserves':
                    # Reserves = Previous + Retained Earnings (PAT * retention)
                    pat_row = self.pnl_rows.get('net_profit')
                    if pat_row:
                        cell.value = f"={prev_col}{row}+'P&L'!{curr_col}{pat_row}*0.70"
                    else:
                        cell.value = f"={prev_col}{row}*1.12"
                        
                elif display_name == 'Total Equity':
                    eq_row = self.bs_rows.get('equity_capital')
                    res_row = self.bs_rows.get('reserves')
                    cell.value = f"={curr_col}{eq_row}+{curr_col}{res_row}"
                    cell.font = Font(bold=True)
                    
                elif display_name == 'Deposits':
                    growth_ref = self.assumption_refs.get('deposit_growth', {}).get(fc_year)
                    if growth_ref:
                        cell.value = f"={prev_col}{row}*(1+{growth_ref}/100)"
                    else:
                        cell.value = f"={prev_col}{row}*1.12"
                        
                elif display_name == 'Borrowings':
                    cell.value = f"={prev_col}{row}*1.08"
                    
                elif display_name == 'Other Liabilities':
                    cell.value = f"={prev_col}{row}*1.10"
                    
                elif display_name == 'Total Liabilities':
                    # Total = Equity + Deposits + Borrowings + Other
                    te_row = self.bs_rows.get('total_equity')
                    dep_row = self.bs_rows.get('deposits')
                    borr_row = self.bs_rows.get('borrowings')
                    oth_row = self.bs_rows.get('other_liabilities')
                    cell.value = f"={curr_col}{te_row}+{curr_col}{dep_row}+{curr_col}{borr_row}+{curr_col}{oth_row}"
                    cell.font = Font(bold=True)
                    
                elif display_name == 'Fixed Assets':
                    cell.value = f"={prev_col}{row}*1.05"
                    
                elif display_name == 'CWIP':
                    cell.value = f"={prev_col}{row}*1.02"
                    
                elif display_name == 'Investments':
                    cell.value = f"={prev_col}{row}*1.10"
                    
                elif display_name == 'Advances':
                    growth_ref = self.assumption_refs.get('loan_growth_rate', {}).get(fc_year)
                    if growth_ref:
                        cell.value = f"={prev_col}{row}*(1+{growth_ref}/100)"
                    else:
                        cell.value = f"={prev_col}{row}*1.12"
                        
                elif display_name == 'Other Assets':
                    cell.value = f"={prev_col}{row}*1.08"
                    
                elif display_name == 'Total Assets':
                    # Total Assets = Total Liabilities (BS identity)
                    tl_row = self.bs_rows.get('total_liabilities')
                    cell.value = f"={curr_col}{tl_row}"
                    cell.font = Font(bold=True)
                
                if display_name not in ['LIABILITIES', 'ASSETS']:
                    cell.number_format = '#,##0.00'
                    cell.fill = FORECAST_FILL
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='right')
            
            row += 1
        
        ws.column_dimensions['A'].width = 22
        for i in range(2, len(all_years) + 2):
            ws.column_dimensions[get_column_letter(i)].width = 14
    
    def _create_roe_tree_sheet(self, data: Dict[str, Any]):
        """Create ROE Decomposition Tree (DuPont Analysis)"""
        ws = self.wb.create_sheet("ROE Tree")
        
        ws['A1'] = "ROE DECOMPOSITION (DUPONT ANALYSIS)"
        ws['A1'].font = TITLE_FONT
        
        ws['A2'] = "ROE = Net Profit Margin x Asset Turnover x Equity Multiplier"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        
        hist_years = self._get_historical_years(data)
        all_years = hist_years + self.forecast_years
        
        # Headers
        ws['A4'] = "Metric"
        for i, year in enumerate(all_years, start=2):
            cell = ws.cell(row=4, column=i)
            cell.value = year
            if year in self.forecast_years:
                cell.fill = FORECAST_FILL
        self._apply_header_style(ws, 4, 1, len(all_years) + 1)
        
        forecast_col_start = len(hist_years) + 2
        
        # ROE Tree structure
        roe_items = [
            ("", "", "section"),
            ("PROFITABILITY", "", "section"),
            ("Net Profit (PAT)", "pat", "link"),
            ("Revenue", "revenue", "link"),
            ("Net Profit Margin (%)", "npm", "calc"),
            ("", "", "spacer"),
            ("EFFICIENCY", "", "section"),
            ("Revenue", "revenue2", "link"),
            ("Average Total Assets", "avg_assets", "link"),
            ("Asset Turnover (x)", "asset_turnover", "calc"),
            ("", "", "spacer"),
            ("LEVERAGE", "", "section"),
            ("Average Total Assets", "avg_assets2", "link"),
            ("Average Equity", "avg_equity", "link"),
            ("Equity Multiplier (x)", "eq_multiplier", "calc"),
            ("", "", "spacer"),
            ("ROE CALCULATION", "", "section"),
            ("ROE (%)", "roe", "final"),
            ("ROE Check: NPM x AT x EM (%)", "roe_check", "verify"),
        ]
        
        row = 5
        roe_rows = {}
        
        for label, key, item_type in roe_items:
            if item_type == "spacer":
                row += 1
                continue
            
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).border = THIN_BORDER
            
            if item_type == "section":
                ws.cell(row=row, column=1).font = Font(bold=True, color="1F4E79")
                ws.cell(row=row, column=1).fill = ROE_FILL
                for col in range(2, len(all_years) + 2):
                    ws.cell(row=row, column=col).fill = ROE_FILL
                    ws.cell(row=row, column=col).border = THIN_BORDER
            else:
                if key:
                    roe_rows[key] = row
                
                for fc_idx, year in enumerate(all_years):
                    col_idx = fc_idx + 2
                    cell = ws.cell(row=row, column=col_idx)
                    curr_col = get_column_letter(col_idx)
                    
                    is_forecast = year in self.forecast_years
                    
                    # Link to P&L and Balance Sheet
                    if key == 'pat':
                        pat_row = self.pnl_rows.get('pat')
                        if pat_row:
                            cell.value = f"='P&L'!{curr_col}{pat_row}"
                            
                    elif key in ['revenue', 'revenue2']:
                        rev_row = self.pnl_rows.get('revenue')
                        if rev_row:
                            cell.value = f"='P&L'!{curr_col}{rev_row}"
                            
                    elif key == 'npm':
                        # Net Profit Margin = PAT / Revenue * 100
                        pat_r = roe_rows.get('pat')
                        rev_r = roe_rows.get('revenue')
                        if pat_r and rev_r:
                            cell.value = f"=IF({curr_col}{rev_r}=0,0,{curr_col}{pat_r}/{curr_col}{rev_r}*100)"
                        cell.number_format = '0.00'
                        cell.font = Font(bold=True)
                        
                    elif key in ['avg_assets', 'avg_assets2']:
                        assets_row = self.bs_rows.get('total_assets')
                        if assets_row and col_idx > 2:
                            prev_col = get_column_letter(col_idx - 1)
                            cell.value = f"=('Balance Sheet'!{curr_col}{assets_row}+'Balance Sheet'!{prev_col}{assets_row})/2"
                        elif assets_row:
                            cell.value = f"='Balance Sheet'!{curr_col}{assets_row}"
                            
                    elif key == 'avg_equity':
                        eq_row = self.bs_rows.get('total_equity')
                        if eq_row and col_idx > 2:
                            prev_col = get_column_letter(col_idx - 1)
                            cell.value = f"=('Balance Sheet'!{curr_col}{eq_row}+'Balance Sheet'!{prev_col}{eq_row})/2"
                        elif eq_row:
                            cell.value = f"='Balance Sheet'!{curr_col}{eq_row}"
                            
                    elif key == 'asset_turnover':
                        # Asset Turnover = Revenue / Avg Assets
                        rev_r = roe_rows.get('revenue2')
                        assets_r = roe_rows.get('avg_assets')
                        if rev_r and assets_r:
                            cell.value = f"=IF({curr_col}{assets_r}=0,0,{curr_col}{rev_r}/{curr_col}{assets_r})"
                        cell.number_format = '0.00'
                        cell.font = Font(bold=True)
                        
                    elif key == 'eq_multiplier':
                        # Equity Multiplier = Avg Assets / Avg Equity
                        assets_r = roe_rows.get('avg_assets2')
                        eq_r = roe_rows.get('avg_equity')
                        if assets_r and eq_r:
                            cell.value = f"=IF({curr_col}{eq_r}=0,0,{curr_col}{assets_r}/{curr_col}{eq_r})"
                        cell.number_format = '0.00'
                        cell.font = Font(bold=True)
                        
                    elif key == 'roe':
                        # ROE = PAT / Avg Equity * 100
                        pat_r = roe_rows.get('pat')
                        eq_r = roe_rows.get('avg_equity')
                        if pat_r and eq_r:
                            cell.value = f"=IF({curr_col}{eq_r}=0,0,{curr_col}{pat_r}/{curr_col}{eq_r}*100)"
                        cell.number_format = '0.00'
                        cell.font = Font(bold=True, size=12, color="1F4E79")
                        
                    elif key == 'roe_check':
                        # ROE Check = NPM * AT * EM (should equal ROE)
                        npm_r = roe_rows.get('npm')
                        at_r = roe_rows.get('asset_turnover')
                        em_r = roe_rows.get('eq_multiplier')
                        if npm_r and at_r and em_r:
                            cell.value = f"={curr_col}{npm_r}/100*{curr_col}{at_r}*{curr_col}{em_r}*100"
                        cell.number_format = '0.00'
                        cell.font = Font(italic=True, color="666666")
                    
                    if is_forecast and item_type != "section":
                        cell.fill = FORECAST_FILL
                    
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(horizontal='right')
                    if not cell.number_format:
                        cell.number_format = '#,##0.00'
            
            row += 1
        
        # Add explanation
        row += 2
        ws.cell(row=row, column=1).value = "INTERPRETATION"
        ws.cell(row=row, column=1).font = SECTION_FONT
        row += 1
        ws.cell(row=row, column=1).value = "ROE = Net Profit Margin x Asset Turnover x Equity Multiplier"
        row += 1
        ws.cell(row=row, column=1).value = "- Net Profit Margin: How much profit per rupee of revenue (profitability)"
        row += 1
        ws.cell(row=row, column=1).value = "- Asset Turnover: How efficiently assets generate revenue (efficiency)"
        row += 1
        ws.cell(row=row, column=1).value = "- Equity Multiplier: Financial leverage (higher = more debt)"
        
        ws.column_dimensions['A'].width = 35
        for i in range(2, len(all_years) + 2):
            ws.column_dimensions[get_column_letter(i)].width = 14
    
    def _create_quarterly_sheet(self, data: Dict[str, Any]):
        """Create Quarterly Results sheet"""
        ws = self.wb.create_sheet("Quarterly")
        
        ws['A1'] = "QUARTERLY RESULTS"
        ws['A1'].font = TITLE_FONT
        
        operational = data.get('operational_data', {})
        quarterly = operational.get('quarterly_results', {})
        
        if quarterly:
            quarters = sorted(quarterly.keys(), key=lambda x: x)[-12:]
        else:
            quarters = []
        
        ws['A3'] = "Line Item"
        for i, q in enumerate(quarters, start=2):
            ws.cell(row=3, column=i).value = q
        if quarters:
            self._apply_header_style(ws, 3, 1, len(quarters) + 1)
        
        row = 4
        if quarterly and quarters:
            first_q_data = quarterly.get(quarters[0], {})
            for item in list(first_q_data.keys()):
                ws.cell(row=row, column=1).value = item
                ws.cell(row=row, column=1).border = THIN_BORDER
                
                for col_idx, q in enumerate(quarters, start=2):
                    q_data = quarterly.get(q, {})
                    value = q_data.get(item)
                    
                    cell = ws.cell(row=row, column=col_idx)
                    if value is not None:
                        cell.value = value
                        cell.number_format = '#,##0.00'
                    cell.border = THIN_BORDER
                    cell.font = NUMBER_FONT
                    cell.alignment = Alignment(horizontal='right')
                
                row += 1
        
        ws.column_dimensions['A'].width = 25
        for i in range(2, max(len(quarters) + 2, 3)):
            ws.column_dimensions[get_column_letter(i)].width = 12
    
    def _create_ratios_sheet(self, data: Dict[str, Any]):
        """Create Key Ratios sheet"""
        ws = self.wb.create_sheet("Key Ratios")
        
        ws['A1'] = "KEY FINANCIAL RATIOS"
        ws['A1'].font = TITLE_FONT
        
        historical = data.get('historical_financials', {})
        ratios = historical.get('ratios', {})
        
        if ratios:
            years = sorted([y for y in ratios.keys() if y != 'TTM'], key=lambda x: x)[-10:]
        else:
            years = []
        
        ws['A3'] = "Ratio"
        for i, year in enumerate(years, start=2):
            ws.cell(row=3, column=i).value = year
        if years:
            self._apply_header_style(ws, 3, 1, len(years) + 1)
        
        row = 4
        if ratios and years:
            first_year_data = ratios.get(years[0], {})
            for item in list(first_year_data.keys()):
                ws.cell(row=row, column=1).value = item
                ws.cell(row=row, column=1).border = THIN_BORDER
                
                for col_idx, year in enumerate(years, start=2):
                    year_data = ratios.get(year, {})
                    value = year_data.get(item)
                    
                    cell = ws.cell(row=row, column=col_idx)
                    if value is not None:
                        cell.value = value
                        cell.number_format = '0.00'
                    cell.border = THIN_BORDER
                    cell.font = NUMBER_FONT
                    cell.alignment = Alignment(horizontal='right')
                
                row += 1
        
        ws.column_dimensions['A'].width = 25
        for i in range(2, max(len(years) + 2, 3)):
            ws.column_dimensions[get_column_letter(i)].width = 12
    
    def _create_valuation_sheet(self, data: Dict[str, Any]):
        """Create Valuation sheet with RIV model - FORMULA DRIVEN linked to projections"""
        ws = self.wb.create_sheet("Valuation")
        
        valuation = data.get('valuation', {})
        metadata = data.get('company_metadata', {})
        
        ws['A1'] = "VALUATION MODEL"
        ws['A1'].font = TITLE_FONT
        
        ws['A2'] = "Residual Income Valuation (RIV) - All values linked to projected financials"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        
        # Store valuation row references
        val_rows = {}
        
        # ===== SECTION 1: INPUT ASSUMPTIONS =====
        ws['A4'] = "VALUATION INPUTS"
        ws['A4'].font = SECTION_FONT
        
        ws['A6'] = "Parameter"
        ws['B6'] = "Value"
        ws['C6'] = "Note"
        self._apply_header_style(ws, 6, 1, 3)
        
        row = 7
        
        # Risk-free rate
        ws.cell(row=row, column=1).value = "Risk-Free Rate (Rf)"
        ws.cell(row=row, column=2).value = 0.07
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=2).fill = ASSUMPTION_FILL
        ws.cell(row=row, column=3).value = "10-year G-Sec yield"
        val_rows['rf'] = row
        row += 1
        
        # Equity risk premium
        ws.cell(row=row, column=1).value = "Equity Risk Premium"
        ws.cell(row=row, column=2).value = 0.06
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=2).fill = ASSUMPTION_FILL
        ws.cell(row=row, column=3).value = "India market premium"
        val_rows['erp'] = row
        row += 1
        
        # Beta
        ws.cell(row=row, column=1).value = "Beta"
        ws.cell(row=row, column=2).value = 1.0
        ws.cell(row=row, column=2).number_format = '0.00'
        ws.cell(row=row, column=2).fill = ASSUMPTION_FILL
        ws.cell(row=row, column=3).value = "Company beta"
        val_rows['beta'] = row
        row += 1
        
        # Cost of equity (FORMULA)
        ws.cell(row=row, column=1).value = "Cost of Equity (Ke)"
        ws.cell(row=row, column=2).value = f"=B{val_rows['rf']}+B{val_rows['beta']}*B{val_rows['erp']}"
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=3).value = "=Rf + Beta × ERP"
        val_rows['ke'] = row
        row += 1
        
        # Terminal growth
        ws.cell(row=row, column=1).value = "Terminal Growth (g)"
        ws.cell(row=row, column=2).value = 0.04
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=2).fill = ASSUMPTION_FILL
        ws.cell(row=row, column=3).value = "Long-term GDP growth"
        val_rows['g'] = row
        row += 1
        
        # Shares outstanding
        shares = metadata.get('shares_outstanding', 1538)  # In crores
        ws.cell(row=row, column=1).value = "Shares Outstanding (Cr)"
        ws.cell(row=row, column=2).value = shares
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        ws.cell(row=row, column=2).fill = ASSUMPTION_FILL
        val_rows['shares'] = row
        row += 2
        
        # ===== SECTION 2: PROJECTED FINANCIALS (Linked to P&L and BS) =====
        ws.cell(row=row, column=1).value = "PROJECTED FINANCIALS"
        ws.cell(row=row, column=1).font = SECTION_FONT
        row += 2
        
        # Headers for projection years
        ws.cell(row=row, column=1).value = "Metric"
        hist_years = self._get_historical_years(data)
        last_hist_col = len(hist_years) + 1
        
        # We'll use forecast years for valuation
        for i, year in enumerate(self.forecast_years, start=2):
            ws.cell(row=row, column=i).value = year
            ws.cell(row=row, column=i).fill = FORECAST_FILL
        self._apply_header_style(ws, row, 1, len(self.forecast_years) + 1)
        row += 1
        header_row = row - 1
        
        # Net Profit (linked to P&L)
        ws.cell(row=row, column=1).value = "Net Profit (Rs. Cr)"
        pat_row = self.pnl_rows.get('pat')
        for i, year in enumerate(self.forecast_years, start=2):
            col_letter = get_column_letter(last_hist_col + i - 1)  # Map to P&L columns
            if pat_row:
                ws.cell(row=row, column=i).value = f"='P&L'!{col_letter}{pat_row}"
            ws.cell(row=row, column=i).number_format = '#,##0'
            ws.cell(row=row, column=i).fill = FORECAST_FILL
        val_rows['pat'] = row
        row += 1
        
        # Book Value / Equity (linked to Balance Sheet)
        ws.cell(row=row, column=1).value = "Shareholders' Equity (Rs. Cr)"
        eq_row = self.bs_rows.get('total_equity')
        for i, year in enumerate(self.forecast_years, start=2):
            col_letter = get_column_letter(last_hist_col + i - 1)
            if eq_row:
                ws.cell(row=row, column=i).value = f"='Balance Sheet'!{col_letter}{eq_row}"
            ws.cell(row=row, column=i).number_format = '#,##0'
            ws.cell(row=row, column=i).fill = FORECAST_FILL
        val_rows['equity'] = row
        row += 1
        
        # Opening Equity (previous year)
        ws.cell(row=row, column=1).value = "Opening Equity (Rs. Cr)"
        for i, year in enumerate(self.forecast_years, start=2):
            if i == 2:
                # First forecast year - use last historical
                if eq_row:
                    ws.cell(row=row, column=i).value = f"='Balance Sheet'!{get_column_letter(last_hist_col)}{eq_row}"
            else:
                # Previous year's closing equity
                prev_col = get_column_letter(i - 1)
                ws.cell(row=row, column=i).value = f"={prev_col}{val_rows['equity']}"
            ws.cell(row=row, column=i).number_format = '#,##0'
            ws.cell(row=row, column=i).fill = FORECAST_FILL
        val_rows['opening_equity'] = row
        row += 1
        
        # Average Equity
        ws.cell(row=row, column=1).value = "Average Equity (Rs. Cr)"
        for i in range(2, len(self.forecast_years) + 2):
            col = get_column_letter(i)
            ws.cell(row=row, column=i).value = f"=({col}{val_rows['equity']}+{col}{val_rows['opening_equity']})/2"
            ws.cell(row=row, column=i).number_format = '#,##0'
            ws.cell(row=row, column=i).fill = FORECAST_FILL
        val_rows['avg_equity'] = row
        row += 1
        
        # ROE (calculated)
        ws.cell(row=row, column=1).value = "ROE (%)"
        ws.cell(row=row, column=1).font = Font(bold=True)
        for i in range(2, len(self.forecast_years) + 2):
            col = get_column_letter(i)
            ws.cell(row=row, column=i).value = f"=IF({col}{val_rows['avg_equity']}=0,0,{col}{val_rows['pat']}/{col}{val_rows['avg_equity']})"
            ws.cell(row=row, column=i).number_format = '0.0%'
            ws.cell(row=row, column=i).font = Font(bold=True)
            ws.cell(row=row, column=i).fill = FORECAST_FILL
        val_rows['roe'] = row
        row += 2
        
        # ===== SECTION 3: RESIDUAL INCOME CALCULATION =====
        ws.cell(row=row, column=1).value = "RESIDUAL INCOME CALCULATION"
        ws.cell(row=row, column=1).font = SECTION_FONT
        row += 2
        
        # Headers again
        ws.cell(row=row, column=1).value = "Component"
        for i, year in enumerate(self.forecast_years, start=2):
            ws.cell(row=row, column=i).value = year
            ws.cell(row=row, column=i).fill = FORECAST_FILL
        self._apply_header_style(ws, row, 1, len(self.forecast_years) + 1)
        row += 1
        
        # Required Return = Opening Equity × Cost of Equity
        ws.cell(row=row, column=1).value = "Required Return (Rs. Cr)"
        for i in range(2, len(self.forecast_years) + 2):
            col = get_column_letter(i)
            ws.cell(row=row, column=i).value = f"={col}{val_rows['opening_equity']}*$B${val_rows['ke']}"
            ws.cell(row=row, column=i).number_format = '#,##0'
            ws.cell(row=row, column=i).fill = FORECAST_FILL
        val_rows['required_return'] = row
        row += 1
        
        # Residual Income = Net Profit - Required Return
        ws.cell(row=row, column=1).value = "Residual Income (Rs. Cr)"
        ws.cell(row=row, column=1).font = Font(bold=True)
        for i in range(2, len(self.forecast_years) + 2):
            col = get_column_letter(i)
            ws.cell(row=row, column=i).value = f"={col}{val_rows['pat']}-{col}{val_rows['required_return']}"
            ws.cell(row=row, column=i).number_format = '#,##0'
            ws.cell(row=row, column=i).font = Font(bold=True)
            ws.cell(row=row, column=i).fill = FORECAST_FILL
        val_rows['residual_income'] = row
        row += 1
        
        # Discount Factor
        ws.cell(row=row, column=1).value = "Discount Factor"
        for i, year_idx in enumerate(range(len(self.forecast_years)), start=2):
            ws.cell(row=row, column=i).value = f"=1/(1+$B${val_rows['ke']})^{year_idx + 1}"
            ws.cell(row=row, column=i).number_format = '0.000'
            ws.cell(row=row, column=i).fill = FORECAST_FILL
        val_rows['discount_factor'] = row
        row += 1
        
        # PV of Residual Income
        ws.cell(row=row, column=1).value = "PV of Residual Income (Rs. Cr)"
        for i in range(2, len(self.forecast_years) + 2):
            col = get_column_letter(i)
            ws.cell(row=row, column=i).value = f"={col}{val_rows['residual_income']}*{col}{val_rows['discount_factor']}"
            ws.cell(row=row, column=i).number_format = '#,##0'
            ws.cell(row=row, column=i).fill = FORECAST_FILL
        val_rows['pv_ri'] = row
        row += 2
        
        # ===== SECTION 4: VALUATION SUMMARY =====
        ws.cell(row=row, column=1).value = "VALUATION SUMMARY"
        ws.cell(row=row, column=1).font = SECTION_FONT
        row += 2
        
        # Current Book Value (from last historical)
        ws.cell(row=row, column=1).value = "Current Book Value (Rs. Cr)"
        if eq_row:
            ws.cell(row=row, column=2).value = f"='Balance Sheet'!{get_column_letter(last_hist_col)}{eq_row}"
        else:
            ws.cell(row=row, column=2).value = metadata.get('book_value', 0) * shares
        ws.cell(row=row, column=2).number_format = '#,##0'
        val_rows['current_bv'] = row
        row += 1
        
        # Sum of PV of Residual Income
        ws.cell(row=row, column=1).value = "Sum of PV of RI (Rs. Cr)"
        pv_cols = [get_column_letter(i) for i in range(2, len(self.forecast_years) + 2)]
        pv_sum_formula = "+".join([f"{col}{val_rows['pv_ri']}" for col in pv_cols])
        ws.cell(row=row, column=2).value = f"={pv_sum_formula}"
        ws.cell(row=row, column=2).number_format = '#,##0'
        val_rows['sum_pv_ri'] = row
        row += 1
        
        # Terminal Value
        last_ri_col = get_column_letter(len(self.forecast_years) + 1)
        ws.cell(row=row, column=1).value = "Terminal Value (Rs. Cr)"
        ws.cell(row=row, column=2).value = f"={last_ri_col}{val_rows['residual_income']}*(1+$B${val_rows['g']})/($B${val_rows['ke']}-$B${val_rows['g']})*{last_ri_col}{val_rows['discount_factor']}"
        ws.cell(row=row, column=2).number_format = '#,##0'
        val_rows['terminal_value'] = row
        row += 1
        
        # Total Equity Value
        ws.cell(row=row, column=1).value = "Total Equity Value (Rs. Cr)"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = f"=B{val_rows['current_bv']}+B{val_rows['sum_pv_ri']}+B{val_rows['terminal_value']}"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=2).font = Font(bold=True)
        val_rows['total_equity_value'] = row
        row += 2
        
        # ===== PER SHARE VALUES =====
        ws.cell(row=row, column=1).value = "PER SHARE VALUATION"
        ws.cell(row=row, column=1).font = SECTION_FONT
        row += 2
        
        # Fair Value per Share
        ws.cell(row=row, column=1).value = "Fair Value per Share (Rs.)"
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row, column=2).value = f"=B{val_rows['total_equity_value']}/B{val_rows['shares']}"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=2).font = Font(bold=True, size=14, color="1F4E79")
        val_rows['fair_value'] = row
        row += 1
        
        # Current Market Price
        cmp = metadata.get('current_price', 0)
        ws.cell(row=row, column=1).value = "Current Market Price (Rs.)"
        ws.cell(row=row, column=2).value = cmp
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        val_rows['cmp'] = row
        row += 1
        
        # Upside/Downside
        ws.cell(row=row, column=1).value = "Upside / Downside (%)"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = f"=(B{val_rows['fair_value']}-B{val_rows['cmp']})/B{val_rows['cmp']}"
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=2).font = Font(bold=True)
        val_rows['upside'] = row
        row += 2
        
        # Recommendation
        ws.cell(row=row, column=1).value = "RECOMMENDATION"
        ws.cell(row=row, column=1).font = SECTION_FONT
        row += 1
        
        rec = valuation.get('recommendation', 'HOLD')
        rec_color = "00AA00" if rec == "BUY" else "CC0000" if rec == "SELL" else "FF9900"
        ws.cell(row=row, column=1).value = rec
        ws.cell(row=row, column=1).font = Font(size=24, bold=True, color=rec_color)
        row += 2
        
        ws.cell(row=row, column=1).value = valuation.get('rationale', 'Based on RIV valuation model')
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        
        # Apply borders to all data cells
        for r in range(6, row):
            for c in range(1, 4):
                ws.cell(row=r, column=c).border = THIN_BORDER
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
    
    def _create_peer_comparison_sheet(self, data: Dict[str, Any]):
        """Create Peer Comparison sheet"""
        ws = self.wb.create_sheet("Peer Comparison")
        
        ws['A1'] = "PEER COMPARISON"
        ws['A1'].font = TITLE_FONT
        
        commentary = data.get('management_commentary', {})
        peers = commentary.get('peer_comparison', [])
        
        if not peers:
            ws['A3'] = "No peer data available"
            return
        
        # Get columns from first peer
        if peers:
            columns = ['name'] + [k for k in peers[0].keys() if k != 'name']
            
            # Headers
            for i, col in enumerate(columns, start=1):
                ws.cell(row=3, column=i).value = col.replace('_', ' ').title()
            self._apply_header_style(ws, 3, 1, len(columns))
            
            # Data
            for row_idx, peer in enumerate(peers, start=4):
                for col_idx, col in enumerate(columns, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = peer.get(col)
                    if value is not None:
                        cell.value = value
                        if isinstance(value, (int, float)):
                            cell.number_format = '#,##0.00'
                    cell.border = THIN_BORDER
                    cell.font = NUMBER_FONT
        
        for i in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 15
    
    def _create_thesis_sheet(self, data: Dict[str, Any]):
        """Create Investment Thesis sheet"""
        ws = self.wb.create_sheet("Thesis")
        
        thesis = data.get('thesis', {})
        valuation = data.get('valuation', {})
        metadata = data.get('company_metadata', {})
        commentary = data.get('management_commentary', {})
        
        ws['A1'] = "INVESTMENT THESIS"
        ws['A1'].font = TITLE_FONT
        
        ws['A2'] = f"{metadata.get('full_name', 'Company')} ({metadata.get('ticker', '')})"
        ws['A2'].font = Font(size=14, bold=True)
        
        # Summary box
        rec = valuation.get('recommendation', 'HOLD')
        ws['A4'] = f"{rec} | Target: Rs.{valuation.get('target_price', 0):,.0f} | Upside: {valuation.get('upside_percent', 0):.1f}%"
        ws['A4'].font = Font(size=12, bold=True)
        
        # Pros section
        row = 7
        ws.cell(row=row, column=1).value = "STRENGTHS (from Screener.in)"
        ws.cell(row=row, column=1).font = SECTION_FONT
        row += 1
        
        pros = commentary.get('pros', [])
        for pro in pros[:5]:
            ws.cell(row=row, column=1).value = f"* {pro}"
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            row += 1
        
        # Cons section
        row += 1
        ws.cell(row=row, column=1).value = "RISKS (from Screener.in)"
        ws.cell(row=row, column=1).font = SECTION_FONT
        row += 1
        
        cons = commentary.get('cons', [])
        for con in cons[:5]:
            ws.cell(row=row, column=1).value = f"* {con}"
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            row += 1
        
        # Full thesis
        row += 2
        ws.cell(row=row, column=1).value = "DETAILED ANALYSIS"
        ws.cell(row=row, column=1).font = SECTION_FONT
        row += 1
        
        thesis_text = thesis.get('full_text', 'See valuation and assumptions sheets for details.')
        ws.cell(row=row, column=1).value = thesis_text
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        
        ws.column_dimensions['A'].width = 100
        ws.row_dimensions[row].height = 300
